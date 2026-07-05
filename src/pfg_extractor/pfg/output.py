from __future__ import annotations
import csv
import logging
import os
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def _ensure_dir(path: str) -> None:
    Path(path).mkdir(parents=True, exist_ok=True)


# ── CSV writer ─────────────────────────────────────────────────────────────────

class CsvWriter:
    """Append-mode CSV writer with header-on-creation and dedup by line_item_id."""

    def __init__(self, filepath: str, columns: list[str]) -> None:
        self._path = Path(filepath)
        self._columns = columns
        self._existing_ids: set[str] = set()
        _ensure_dir(str(self._path.parent))
        self._init()

    def _init(self) -> None:
        if self._path.exists():
            self._load_existing_ids()
        else:
            with open(self._path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.DictWriter(fh, fieldnames=self._columns, extrasaction="ignore")
                writer.writeheader()

    def _load_existing_ids(self) -> None:
        try:
            with open(self._path, newline="", encoding="utf-8") as fh:
                reader = csv.DictReader(fh)
                for row in reader:
                    rid = row.get("line_item_id", "")
                    if rid:
                        self._existing_ids.add(rid)
        except Exception as exc:
            logger.warning("Could not read existing CSV for dedup: %s", exc)

    def write_rows(self, rows: list[dict], dry_run: bool = False) -> int:
        """Write rows that are not already in the CSV. Returns count written."""
        new_rows = [r for r in rows if r.get("line_item_id") not in self._existing_ids]
        if not new_rows:
            return 0
        if dry_run:
            logger.info("[dry-run] Would write %d rows to %s", len(new_rows), self._path)
            return len(new_rows)

        with open(self._path, "a", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self._columns, extrasaction="ignore")
            for row in new_rows:
                writer.writerow(row)
                self._existing_ids.add(row["line_item_id"])

        return len(new_rows)

    @property
    def known_ids(self) -> set:
        return self._existing_ids


# ── XLSX writer ───────────────────────────────────────────────────────────────

def refresh_xlsx(csv_path: str, xlsx_path: str, columns: list[str]) -> None:
    """Rebuild the XLSX from the current CSV (full refresh, not incremental)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed — skipping XLSX output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PFG Line Items"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write rows from CSV
    try:
        import csv as csv_mod
        with open(csv_path, newline="", encoding="utf-8") as fh:
            reader = csv_mod.DictReader(fh)
            for row_idx, row in enumerate(reader, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = row.get(col_name, "")
                    # Coerce numeric-looking strings
                    if val in ("True", "False"):
                        val = val == "True"
                    elif val != "":
                        try:
                            val = int(val)
                        except ValueError:
                            try:
                                val = float(val)
                            except ValueError:
                                pass
                    ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
    except Exception as exc:
        logger.error("Error reading CSV for XLSX: %s", exc)

    # Auto-size columns
    for col_idx, col_name in enumerate(columns, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(len(col_name) + 2, 12)

    wb.save(xlsx_path)
    logger.info("XLSX written to %s", xlsx_path)


# ── Error log ─────────────────────────────────────────────────────────────────

class ErrorLog:
    """Append-mode CSV for parse errors and unparsed lines."""

    COLUMNS = ["message_id", "order_number", "error_type", "raw_text"]

    def __init__(self, filepath: str) -> None:
        self._path = Path(filepath)
        _ensure_dir(str(self._path.parent))
        if not self._path.exists():
            with open(self._path, "w", newline="", encoding="utf-8") as fh:
                csv.DictWriter(fh, fieldnames=self.COLUMNS).writeheader()

    def write(self, message_id: str, order_number: str, error_type: str, raw_text: str) -> None:
        with open(self._path, "a", newline="", encoding="utf-8") as fh:
            csv.DictWriter(fh, fieldnames=self.COLUMNS, extrasaction="ignore").writerow(
                {
                    "message_id": message_id,
                    "order_number": order_number,
                    "error_type": error_type,
                    "raw_text": raw_text[:500],
                }
            )
